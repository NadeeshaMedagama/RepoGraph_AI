# ═══════════════════════════════════════════════════════════════════════════════
# RepoGraph AI - Spreadsheet Processor
# ═══════════════════════════════════════════════════════════════════════════════
# Processor for spreadsheet files (XLSX, XLS)
# ═══════════════════════════════════════════════════════════════════════════════

from typing import Set, List, Any

from models import Document, FileCategory, FileType, ProcessingStatus
from processors.base_processor import BaseProcessor
from utils import get_logger

logger = get_logger(__name__)


class SpreadsheetProcessor(BaseProcessor):
    """
    Processor for spreadsheet files.

    Handles XLSX and XLS files.
    Extracts data from sheets including headers and values.
    """

    @property
    def supported_types(self) -> Set[FileType]:
        return {
            FileType.XLSX,
            FileType.XLS,
        }

    @property
    def category(self) -> FileCategory:
        return FileCategory.SPREADSHEET

    def extract(self, document: Document) -> Document:
        """
        Extract content from a spreadsheet file.

        Args:
            document: Document to process

        Returns:
            Document with extracted_text populated
        """
        try:
            document.status = ProcessingStatus.EXTRACTING

            if document.metadata.file_type == FileType.XLSX:
                content = self._extract_xlsx(document)
            else:
                content = self._extract_xls(document)

            document.extracted_text = content
            document.status = ProcessingStatus.COMPLETED

            logger.debug(
                "Spreadsheet processed",
                file=document.metadata.name,
                text_length=len(content),
            )

            return document

        except Exception as e:
            document.status = ProcessingStatus.FAILED
            document.error_message = str(e)
            logger.error(f"Failed to process spreadsheet: {e}")
            return document

    def _extract_xlsx(self, document: Document) -> str:
        """Extract data from XLSX file."""
        try:
            from openpyxl import load_workbook

            wb = load_workbook(document.metadata.path, read_only=True, data_only=True)
            texts = [f"Spreadsheet: {document.metadata.name}"]
            texts.append(f"Sheets: {len(wb.sheetnames)}\n")

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                texts.append(f"\n=== Sheet: {sheet_name} ===")

                # Get dimensions
                if sheet.max_row and sheet.max_column:
                    texts.append(f"Size: {sheet.max_row} rows x {sheet.max_column} columns\n")

                # Extract data (limit to first 100 rows for large sheets)
                row_limit = min(sheet.max_row or 0, 100)

                rows_data = []
                for row in sheet.iter_rows(max_row=row_limit, values_only=True):
                    # Filter empty rows
                    if any(cell is not None for cell in row):
                        row_text = self._format_row(row)
                        rows_data.append(row_text)

                if rows_data:
                    # First row is often headers
                    if len(rows_data) > 1:
                        texts.append(f"Headers: {rows_data[0]}")
                        texts.append("---")
                        texts.extend(rows_data[1:])
                    else:
                        texts.extend(rows_data)

                if row_limit < (sheet.max_row or 0):
                    texts.append(f"\n... ({sheet.max_row - row_limit} more rows)")

            wb.close()
            return "\n".join(texts)

        except Exception as e:
            logger.warning(f"Failed to extract XLSX: {e}")
            return f"Spreadsheet: {document.metadata.name}\n(Content extraction failed)"

    def _extract_xls(self, document: Document) -> str:
        """Extract data from XLS file."""
        try:
            import xlrd

            wb = xlrd.open_workbook(document.metadata.path)
            texts = [f"Spreadsheet: {document.metadata.name}"]
            texts.append(f"Sheets: {len(wb.sheet_names())}\n")

            for sheet_name in wb.sheet_names():
                sheet = wb.sheet_by_name(sheet_name)
                texts.append(f"\n=== Sheet: {sheet_name} ===")
                texts.append(f"Size: {sheet.nrows} rows x {sheet.ncols} columns\n")

                # Extract data (limit to first 100 rows)
                row_limit = min(sheet.nrows, 100)

                for row_idx in range(row_limit):
                    row = [sheet.cell_value(row_idx, col_idx)
                           for col_idx in range(sheet.ncols)]

                    if any(cell for cell in row):
                        row_text = self._format_row(row)
                        texts.append(row_text)

                if row_limit < sheet.nrows:
                    texts.append(f"\n... ({sheet.nrows - row_limit} more rows)")

            return "\n".join(texts)

        except Exception as e:
            logger.warning(f"Failed to extract XLS: {e}")
            return f"Spreadsheet: {document.metadata.name}\n(Content extraction failed)"

    def _format_row(self, row: tuple) -> str:
        """Format a spreadsheet row as text."""
        formatted_cells = []
        for cell in row:
            if cell is None:
                formatted_cells.append("")
            elif isinstance(cell, float):
                # Check if it's an integer
                if cell == int(cell):
                    formatted_cells.append(str(int(cell)))
                else:
                    formatted_cells.append(f"{cell:.2f}")
            else:
                formatted_cells.append(str(cell).strip())

        return " | ".join(formatted_cells)
